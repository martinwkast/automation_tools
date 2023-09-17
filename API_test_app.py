import requests
from openpyxl import Workbook

API_ENDPOINT = "https://jsonplaceholder.typicode.com/todos"

# Fetch data from the API
response = requests.get(API_ENDPOINT)
if response.status_code != 200:
    raise ValueError("Failed to fetch data from the API")
data = response.json()

# Create a new Excel workbook and get the active sheet
wb = Workbook()
ws = wb.active

# Write headers to the first row
headers = ["UserID", "ID", "Title", "Completed"]
for col_num, header in enumerate(headers, 1):
    ws.cell(row=1, column=col_num, value=header)

# Write data to the spreadsheet
for row_num, entry in enumerate(data, 2):
    ws.cell(row=row_num, column=1, value=entry["userId"])
    ws.cell(row=row_num, column=2, value=entry["id"])
    ws.cell(row=row_num, column=3, value=entry["title"])
    ws.cell(row=row_num, column=4, value=entry["completed"])

# Save the workbook
wb.save("api_data.xlsx")

print("Data written to api_data.xlsx")
